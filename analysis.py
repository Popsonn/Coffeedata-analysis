#imported my libraries
import mysql.connector
import pandas as pd
import openpyxl
from openpyxl import Workbook

# Connected to my database
mydb = mysql.connector.connect(
  host="127.0.0.1",
  user="root",
  password="popson8448",
  database="coffeedata"
)

# Create a cursor
cursor = mydb.cursor()

def execute_query(query, params=None):
    cursor.execute(query, params)
    return cursor.fetchall()

#Task1 query
task1_query = """
SELECT 
i.item_id, 
i.item_name, 
SUM(o.quantity) AS numberSold, 
SUM(o.quantity * i.item_price) AS revenue,
COUNT(i.item_id) AS totalOrders
FROM orders o
JOIN items i ON o.item_id = i.item_id
GROUP BY i.item_id, i.item_name
ORDER BY revenue DESC;
"""
task1_results = execute_query(task1_query)

#Task2 query
task2_query = """
WITH revenue_by_item AS 
(
SELECT i.item_name, SUM(o.quantity * i.item_price) AS revenue
FROM orders o
JOIN items i ON o.item_id = i.item_id
GROUP BY i.item_name
)
SELECT
i.item_name,
rbi.revenue, 
r.recipe_id,
SUM(r.quantity * ing.ing_price) AS production_cost,
rbi.revenue - SUM(r.quantity * ing.ing_price) AS profit,
(rbi.revenue - SUM(r.quantity * ing.ing_price)) / rbi.revenue * 100 AS profitPercentage
FROM items i
JOIN revenue_by_item rbi ON i.item_name = rbi.item_name 
JOIN recipes r ON i.sku = r.recipe_id
JOIN ingredients ing ON r.ing_id = ing.ing_id
WHERE i.item_name IN ('Cappuccino', 'Latte', 'Flat White', 'Caramel Macchiato', 
'Espresso', 'Mocha', 'White Mocha', 'Cold Coffee', 'Cold Mocha')
GROUP BY
i.item_name, r.recipe_id;
"""
task2_results = execute_query(task2_query)

#Task3 query
task3_query = """
SELECT
DATE_FORMAT(o.created_at, '%Y-%m-%d %H:00') AS hour,
COUNT(DISTINCT o.order_id) AS numberOfOrders,
SUM(o.quantity * i.item_price) AS sales,
SUM(o.quantity * i.item_price) - SUM(r.quantity * ing.ing_price) AS profit
FROM orders o
JOIN items i ON o.item_id = i.item_id
JOIN recipes r ON i.sku = r.recipe_id
JOIN ingredients ing ON r.ing_id = ing.ing_id
GROUP BY hour
ORDER BY hour;
"""
task3_results = execute_query(task3_query)

#Task4 query
task4_query = """
SELECT 
CONCAT(s.first_name, ' ', s.last_name) AS staff_name,
SUM(TIME_TO_SEC(TIMEDIFF(sh.end_time, sh.start_time)) / 3600) AS hoursWorked,
SUM(TIME_TO_SEC(TIMEDIFF(sh.end_time, sh.start_time)) / 3600) * s.sal_per_hour AS totalSalaryEarned
FROM staff s
JOIN rota r ON s.staff_id = r.staff_id
JOIN shift sh ON r.shift_id = sh.shift_id
GROUP BY staff_name,s.sal_per_hour
ORDER BY hoursWorked DESC;
"""
task4_results = execute_query(task4_query)

#Task5 query
task5_query = """
SELECT
oi.item_cat,
CASE WHEN oi.dine_in_order = 1 THEN 'dine-in'
WHEN oi.takeout_order = 1 THEN 'takeout'
END AS inOrOut,
SUM(profit) AS aggregateProfit
FROM
    (SELECT
        o.order_id,
        o.created_at,
        i.item_id,
        o.quantity,
        i.item_price,
        i.item_cat,
        (i.item_price - SUM(r.quantity * ing.ing_price)) * o.quantity AS profit,
        CASE WHEN o.in_or_out = 'in' THEN 1 ELSE 0 END AS dine_in_order,
        CASE WHEN o.in_or_out = 'out' THEN 1 ELSE 0 END AS takeout_order
    FROM
        coffeedata.orders o
        JOIN coffeedata.items i ON o.item_id = i.item_id
        JOIN coffeedata.recipes r ON i.sku = r.recipe_id
        JOIN coffeedata.ingredients ing ON r.ing_id = ing.ing_id
    GROUP BY
        o.order_id, o.created_at, i.item_id, o.quantity, i.item_price, i.item_cat, o.in_or_out
    ) AS oi
GROUP BY oi.item_cat,
inOrOut;
"""
task5_results = execute_query(task5_query)

#Task6 query
task6_query = """
SELECT 
DATE(created_at) AS order_date,
SUM(CASE WHEN HOUR(created_at) BETWEEN 7 AND 13 THEN 1 ELSE 0 END) AS morningShift,
SUM(CASE WHEN HOUR(created_at) BETWEEN 13 AND 17 THEN 1 ELSE 0 END) AS afternoonShift
FROM coffeedata.orders
GROUP BY DATE(created_at);
"""
task6_results = execute_query(task6_query)

# I created a new workbook
workbook = openpyxl.Workbook()

# I remove the default blank sheet named 'Sheet' created automically
default_sheet = workbook["Sheet"]
workbook.remove(default_sheet)

# writing Task 1 results to excel
worksheet = workbook.create_sheet("Task1Results")
worksheet.append(["itemId", "itemName", "numberSold", "revenue", "totalOrders"])
for row in task1_results:
    worksheet.append(list(row))

# writing Task 2 results to excel
worksheet = workbook.create_sheet("Task2Results")
worksheet.append(["itemName", "revenue", "recipeId", "productionCost", "profitOrLoss", "profitOrLossPercentage"])
for row in task2_results:worksheet.append(list(row))

# writing Task 3 results to excel
worksheet = workbook.create_sheet("Task3Results")
worksheet.append(["hour", "numberOfOrders", "sales", "profitOrLoss"])
for row in task3_results:
    worksheet.append(list(row))

# writing Task 4 results to excel
worksheet = workbook.create_sheet("Task4Results")
worksheet.append(["staffName", "hoursWorked", "totalSalaryEarned"])
for row in task4_results:
    worksheet.append(list(row))

# writing Task 5 results to excel
worksheet = workbook.create_sheet("Task5Results")
worksheet.append(["itemCat", "inOrOut", "aggregateProfitOrLoss"])
for row in task5_results:
    worksheet.append(list(row))

# writing Task 6 results to excel
worksheet = workbook.create_sheet("Task6Result")
worksheet.append(["orderDate", "morningShift", "afternoonShift"])
for row in task6_results:
    worksheet.append(list(row))

# Saved the workbook
workbook.save("coffeeAnalysisResults.xlsx")

